from jira import JIRA
from openpyxl.styles import Font
from openpyxl import load_workbook


print('Program starts pulling data from Jira')
# Settings
email = 'jacek.dar.ostaszewski@dxc.com'  # Jira username
api_token = "MTM1Nzg5MDc2MTA0OhgsxPC1KTg+yyMPbBckaBaRNyHi"  # Jira API token
server = 'https://jira.dxc.com'  # Jira server URL
jql = "project = APOMON"  # JQL

# Get issues from Jira in JSON format
jira = JIRA(options={'server': server}, token_auth=(api_token))
# Create all issues for pulling consecutive fields for printing
jira_issues = jira.search_issues(jql, maxResults=0)

# Write starting data where data to be writen in excel
row=0
startcolumn=1
startrow=4

def utworz_liste(issue):
    # utwórz listę do druku z pól issue
    issuekey = str(issue)
    issuetype = jira.issue(issuekey).fields.issuetype
    summary = jira.issue(issuekey).fields.summary
    assignee = jira.issue(issuekey).fields.assignee
    priority = jira.issue(issuekey).fields.priority
    status = jira.issue(issuekey).fields.status
    created = jira.issue(issuekey).fields.created
    duedate = jira.issue(issuekey).fields.duedate
    startdate = jira.issue(issuekey).fields.customfield_15892
    enddate = jira.issue(issuekey).fields.customfield_15893
    plannedclosingdate = jira.issue(issuekey).fields.customfield_10310
    lista_pol = [issuekey, issuetype, summary, assignee, priority, status, created, duedate, startdate, enddate, plannedclosingdate]
    return lista_pol

def write_columntitels(startrow,startcolumn,work_sheet):
    filedslist = ['Issuekey','Issuetype','Summary','Assignee','Priority','Status','Created','Duedate','Startdate','Enddate','PlannedCloseDate']
    for i in range(len(filedslist)):
        mycell = work_sheet.cell(row=startrow,column=startcolumn+i+1)
        mycell.value = filedslist[i]


def drukuj_subtasks(issuesubtasks,work_sheet):
    global row, startrow, startcolumn
    for e in issuesubtasks:
        if e != []:
            print('subtasks: ' + str(e))
            lista_pol_subtask = utworz_liste(e)
            print(*lista_pol_subtask, sep=' ')
            for i in range(len(lista_pol_subtask)):
                mycell = work_sheet.cell(row=row+1+startrow,column=i+1+startcolumn)
                mycell.value = str(lista_pol_subtask[i])
            row = row + 1

def write_to_file(lista_pol,work_sheet,bold):
    global row, startrow, startcolumn
    for col in range(0,len(lista_pol)):
        mycell = work_sheet.cell(row=(1+row+startrow),column=(col+1+startcolumn))
        mycell.value = str(lista_pol[col])
        mycell.font = Font(bold=bold)
    row +=1
print('Now I am popultating tabels with data')
print('Epics in process')
# Definition of table for issues excluding apics and subtasks
tabela = []
# Table for subtasks
epicsubtable = [] # table for subtasks in epic
issuesubtable = [] # table for subtasks in issues


# Creat list of Epics, Issue in Epic and subtasks in Epic
epic_lists =jira.search_issues("'issuetype'=Epic AND project='APOMON'")
# Create list of issues in Epics plus subtasks for Epics in separate subtasks table
for issue in epic_lists:
    issue_list = jira.search_issues("'Epic Link'=%s AND project=%s" %(str(issue), 'APOMON'), maxResults=1000)
    tabela.append(issue_list)
    subtasks = jira.issue(issue).get_field('subtasks')
    epicsubtable.append(subtasks)
print('Issues and subtasks in progress')

# Create list of subtasks for Issues in table: tabela
for i in range(len(tabela)):
    for il in range(len(tabela[i])):
        subtasks = jira.issue(tabela[i][il]).get_field('subtasks')
        issuesubtable.append(subtasks)

print('Start writing data')
subindex = 0
# Open workbook
wb = load_workbook('testfile.xlsx')
work_sheet = wb.active

# Write all the fields for epic and epic subtasks
write_columntitels(startrow,startcolumn,work_sheet)
for a in range(len(tabela)):
    epic = str(epic_lists[a])
    print('Epic:' + epic)
    lista_pol_epica = utworz_liste(epic)
    print(*lista_pol_epica, sep=' ')
 #   for col in range(0,len(lista_pol_epica)):
 #       mycell = work_sheet.cell(row=(1+row),column=(col+1))
 #       mycell.value = str(lista_pol_epica[col])
    epicsubtasks = epicsubtable[a]
 # check whether the issue comprises subtasks or issues, if yes write epic in bold format
    if epicsubtable[a] != [] or tabela[a] != []:
        bold=True
    else: bold=False
    write_to_file(lista_pol_epica,work_sheet,bold)
    drukuj_subtasks(epicsubtasks,work_sheet)

    # Write all the fields for issues and issues subtasks in epics
    for b in range(len(tabela[a])):
        print('Issue: ' + str(tabela[a][b]))
        #print('issue subtasks: ' + str(issuesubtable[subindex])) skasuj jeśli subtaski są dobrze wyświetlane
        #write on terminal fields of issues
        issue = str(tabela[a][b])
        lista_pol_issue = utworz_liste(tabela[a][b])
        print(*lista_pol_issue, sep=' ')

        #write on terminal fields of subtasks - nit needed as there's drukj_subtasks function
        #if issuesubtable[subindex] != []:
        #    for s in range(len(issuesubtable[subindex])):
        #        subtask = str(issuesubtable[subindex][s])
        #        list_pol_subtask = utworz_liste(issuesubtable[subindex][s])
        #        print(*list_pol_subtask, sep=' ')

        #check if issue has underlying subtasks, if yes, change to bold format
        if issuesubtable[subindex] != []:
            bold=True
        else:
            bold=False
        #write issus fields and suntasks fields in excel file
        write_to_file(lista_pol_issue,work_sheet,bold)
        drukuj_subtasks(issuesubtable[subindex],work_sheet)
        subindex = subindex + 1


# Close the exel file
wb.save('testfile.xlsx')
print('The end of the program, check the excel file testfile.xlsx')

